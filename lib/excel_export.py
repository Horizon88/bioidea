"""Excel exporter — render the bioidea SQLite + run artifacts as a single
XLSX file with one tab ("Ideas") and one row per run.

Designed for at-a-glance scanning, not deep editing. Hyperlinks to local
markdown files let the user click into full output. Auto-runs after every
stage completes (with `--no-export` opt-out) and is idempotent — calling it
again rebuilds the entire workbook from current DB + disk.
"""
from __future__ import annotations

import json
import shutil
import sqlite3
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from . import extract as _extract


# Column layout: (key, header, width). Order matters — first-to-last is left-to-right.
COLUMNS: list[tuple[str, str, int]] = [
    ("run_short", "Run", 10),
    ("name", "Name", 32),
    ("status", "Status", 11),
    ("created", "Created", 17),
    ("current_stage", "Cur", 5),
    ("total_cost", "Total $", 9),

    # Stage 0
    ("stage0_target", "S0 Target", 38),
    ("stage0_LGS", "LGS", 5),
    ("stage0_pain_point_$", "S0 TAM ($)", 12),
    ("stage0_qa_decision", "S0 QA Decision", 28),

    # Stage 1
    ("stage1_fto_verdict", "S1 FTO", 22),
    ("stage1_blocker", "S1 Blocker", 22),
    ("stage1_qa_litigation_risk", "Litig", 6),
    ("stage1_qa_verdict", "S1 QA Verdict", 24),

    # Stage 2
    ("stage2_decision", "S2 Decision", 18),
    ("stage2_qa_verdict", "S2 QA Verdict", 22),

    # Stage 3
    ("stage3_throughput", "S3 N-strategy", 12),
    ("stage3_campaign_name", "S3 Campaign", 22),
    ("stage3_qa_verdict", "S3 QA Verdict", 22),

    # Stage 4
    ("stage4_qa_verdict", "S4 QA Verdict", 22),

    # Stage 5
    ("stage5_budget_$", "S5 Budget", 11),
    ("stage5_timeline_weeks", "Wks", 6),
    ("stage5_qa_verdict", "S5 QA Verdict", 22),

    # Stage 6 (the headline)
    ("stage6_code_name", "Code", 12),
    ("stage6_ask_$", "Ask", 10),
    ("stage6_recommendation", "Memo", 6),
    ("stage6_qa_vote", "VC Vote", 16),
    ("stage6_inflection", "Inflection", 50),
    ("stage6_qa_one_thing", "The One Thing", 60),

    # Grounding (lightweight tag-based, from stage outputs)
    ("grounding_status", "Ground", 8),
    ("grounding_pct_tagged", "%Tagged", 8),
    ("grounding_strong", "Strong", 7),
    ("grounding_weak", "Weak", 6),

    # Heavyweight grounded artifacts (from `bioidea ground` -> harness CitationAgent)
    ("grounded_stages", "GR Stages", 11),
    ("grounded_pass_claims", "GR Claims", 10),
    ("grounded_sources", "GR Sources", 11),

    # Stage 0 extra (less prominent)
    ("stage0_thesis", "Hunter's Thesis", 50),

    # Run dir last
    ("run_dir", "Dir", 28),
]


# ---------- styles ----------

HEADER_FILL = PatternFill("solid", start_color="1F2937")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=False)

BODY_FONT = Font(name="Arial", size=10)
BODY_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)

# Status / verdict color map (background fill on the cell)
COLOR_FILLS = {
    "FUND": PatternFill("solid", start_color="DCFCE7"),  # green-100
    "PASS": PatternFill("solid", start_color="FEE2E2"),  # red-100
    "YES": PatternFill("solid", start_color="DCFCE7"),
    "NO": PatternFill("solid", start_color="FEE2E2"),
    "ADVANCE": PatternFill("solid", start_color="DCFCE7"),
    "TERMINATE": PatternFill("solid", start_color="FEE2E2"),
    "CLEAR": PatternFill("solid", start_color="DCFCE7"),
    "CAUTION": PatternFill("solid", start_color="FEF3C7"),  # amber-100
    "BLOCKED": PatternFill("solid", start_color="FEE2E2"),
    "PASS_GR": PatternFill("solid", start_color="DCFCE7"),
    "WARN_GR": PatternFill("solid", start_color="FEF3C7"),
    "FAIL_GR": PatternFill("solid", start_color="FEE2E2"),
    "ROBUST": PatternFill("solid", start_color="DCFCE7"),
    "REVISE": PatternFill("solid", start_color="FEF3C7"),
    "ABORT": PatternFill("solid", start_color="FEE2E2"),
    "PROCEED": PatternFill("solid", start_color="DCFCE7"),
    "BRIDGE": PatternFill("solid", start_color="FEF3C7"),
    "HARD STOP": PatternFill("solid", start_color="FEE2E2"),
    "SAFE TO PROCEED": PatternFill("solid", start_color="DCFCE7"),
    "OPERATIONAL PLAN APPROVED": PatternFill("solid", start_color="DCFCE7"),
    "REVISE TIMELINE": PatternFill("solid", start_color="FEF3C7"),
    "APPROVE": PatternFill("solid", start_color="DCFCE7"),
    "RECOMMEND CALIBRATION": PatternFill("solid", start_color="FEF3C7"),
    "RECOMMEND REVISION": PatternFill("solid", start_color="FEF3C7"),
}


def _color_for(value: str) -> PatternFill | None:
    if not value:
        return None
    v = value.upper().strip()
    # Grounding status
    if v == "PASS":
        # context-sensitive: handled by caller for grounding column
        return None
    # match prefixes — most-specific first
    for key, fill in [
        ("HARD STOP", COLOR_FILLS["HARD STOP"]),
        ("SAFE TO PROCEED", COLOR_FILLS["SAFE TO PROCEED"]),
        ("OPERATIONAL PLAN APPROVED", COLOR_FILLS["OPERATIONAL PLAN APPROVED"]),
        ("REVISE TIMELINE", COLOR_FILLS["REVISE TIMELINE"]),
        ("RECOMMEND CALIBRATION", COLOR_FILLS["RECOMMEND CALIBRATION"]),
        ("RECOMMEND REVISION", COLOR_FILLS["RECOMMEND REVISION"]),
        ("BRIDGE", COLOR_FILLS["BRIDGE"]),
        ("ADVANCE", COLOR_FILLS["ADVANCE"]),
        ("TERMINATE", COLOR_FILLS["TERMINATE"]),
        ("CAUTION", COLOR_FILLS["CAUTION"]),
        ("BLOCKED", COLOR_FILLS["BLOCKED"]),
        ("ROBUST", COLOR_FILLS["ROBUST"]),
        ("REVISE", COLOR_FILLS["REVISE"]),
        ("ABORT", COLOR_FILLS["ABORT"]),
        ("PROCEED", COLOR_FILLS["PROCEED"]),
        ("APPROVE", COLOR_FILLS["APPROVE"]),
        ("CLEAR", COLOR_FILLS["CLEAR"]),
    ]:
        if v.startswith(key):
            return fill
    if v in ("FUND", "YES", "YES — RELEASE FUNDS"):
        return COLOR_FILLS["FUND"]
    if v in ("PASS", "NO", "NO — KILL PROJECT"):
        return COLOR_FILLS["PASS"]
    return None


def _grounding_fill(status: str) -> PatternFill | None:
    if status == "PASS":
        return COLOR_FILLS["PASS_GR"]
    if status == "WARN":
        return COLOR_FILLS["WARN_GR"]
    if status == "FAIL":
        return COLOR_FILLS["FAIL_GR"]
    return None


# ---------- data fetch ----------

def _fetch_runs(db_path: Path) -> list[dict]:
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    runs = conn.execute(
        "SELECT id, name, slug, created_at, focus_profile, current_stage, status "
        "FROM runs ORDER BY created_at DESC"
    ).fetchall()
    out = []
    for r in runs:
        execs = conn.execute(
            "SELECT stage, phase, output_path, cost_usd, "
            "grounding_status, grounding_tagged, grounding_untagged, "
            "grounding_strong, grounding_weak "
            "FROM stage_executions WHERE run_id = ? ORDER BY stage, phase",
            (r["id"],),
        ).fetchall()
        out.append({"run": dict(r), "execs": [dict(e) for e in execs]})
    conn.close()
    return out


def _aggregate_grounding(execs: list[dict]) -> dict:
    statuses = [e["grounding_status"] for e in execs if e.get("grounding_status")]
    if not statuses:
        return {"grounding_status": "", "grounding_pct_tagged": "",
                "grounding_strong": 0, "grounding_weak": 0}
    # worst-of: FAIL > WARN > PASS
    rank = {"FAIL": 2, "WARN": 1, "PASS": 0}
    worst = max(statuses, key=lambda s: rank.get(s, 0))
    tagged = sum((e.get("grounding_tagged") or 0) for e in execs)
    untagged = sum((e.get("grounding_untagged") or 0) for e in execs)
    pct = (tagged / (tagged + untagged) * 100) if (tagged + untagged) > 0 else 0
    strong = sum((e.get("grounding_strong") or 0) for e in execs)
    weak = sum((e.get("grounding_weak") or 0) for e in execs)
    return {
        "grounding_status": worst,
        "grounding_pct_tagged": f"{pct:.0f}%",
        "grounding_strong": strong,
        "grounding_weak": weak,
    }


def _scan_grounded(run_id_short: str, slug: str, repo_root: Path) -> dict:
    """Scan runs/<id>/stage{N}_grounded.json files (produced by `bioidea ground`)
    and aggregate counts."""
    base = repo_root / "runs" / f"{run_id_short}_{slug}"
    out = {"grounded_stages": "", "grounded_pass_claims": 0, "grounded_sources": 0}
    if not base.exists():
        return out
    pass_stages = []
    fail_stages = []
    total_claims = 0
    total_sources = 0
    for p in sorted(base.glob("stage*_grounded.json")):
        try:
            data = json.loads(p.read_text())
        except Exception:
            continue
        # Recover stage number from filename
        try:
            stage_n = int(p.stem.split("_")[0].replace("stage", ""))
        except Exception:
            stage_n = -1
        status = data.get("status", "?")
        if status == "PASS":
            pass_stages.append(str(stage_n))
            total_claims += len(data.get("claims") or [])
            total_sources += len(data.get("sources") or {})
        else:
            fail_stages.append(f"{stage_n}!")
    parts = pass_stages + fail_stages
    out["grounded_stages"] = ",".join(parts)
    out["grounded_pass_claims"] = total_claims
    out["grounded_sources"] = total_sources
    return out


def _build_row(record: dict, repo_root: Path) -> dict:
    run = record["run"]
    execs = record["execs"]

    total_cost = sum(e.get("cost_usd") or 0 for e in execs)

    # collect stage paths for extractors
    stage_paths: dict[tuple[int, str], Path] = {}
    for e in execs:
        if e.get("output_path"):
            stage_paths[(e["stage"], e["phase"])] = Path(e["output_path"])

    fields = _extract.extract_run(stage_paths)
    grounding = _aggregate_grounding(execs)
    grounded = _scan_grounded(run["id"][:8], run["slug"], repo_root)

    run_dir = ""
    if execs and execs[0].get("output_path"):
        try:
            run_dir = str(Path(execs[0]["output_path"]).parent.relative_to(repo_root))
        except Exception:
            run_dir = str(Path(execs[0]["output_path"]).parent)

    return {
        "run_short": run["id"][:8],
        "name": run["name"],
        "status": run["status"],
        "created": run["created_at"][:16].replace("T", " "),
        "current_stage": run["current_stage"],
        "total_cost": f"${total_cost:.2f}",
        "run_dir": run_dir,
        **fields,
        **grounding,
        **grounded,
    }


# ---------- workbook construction ----------

def export_to_xlsx(repo_root: Path, db_path: Path, out_path: Path) -> Path:
    """Build the xlsx and atomically replace `out_path`."""
    records = _fetch_runs(db_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Ideas"
    ws.freeze_panes = "B2"

    # header row
    for col_idx, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 24

    # body rows
    for row_idx, rec in enumerate(records, start=2):
        row = _build_row(rec, repo_root)
        for col_idx, (key, _, _) in enumerate(COLUMNS, start=1):
            value = row.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.alignment = BODY_ALIGN
            cell.border = THIN_BORDER

            # color rules
            if key == "grounding_status":
                fill = _grounding_fill(str(value))
                if fill:
                    cell.fill = fill
            else:
                fill = _color_for(str(value))
                if fill:
                    cell.fill = fill

            # hyperlink the run_dir cell to the local folder
            if key == "run_dir" and value:
                target = repo_root / value
                cell.hyperlink = f"file://{target}"
                cell.font = Font(name="Arial", size=10, color="2563EB", underline="single")

        # row height — give each row ~3 lines for the wrap-friendly text columns
        ws.row_dimensions[row_idx].height = 60

    # write atomically
    out_path.parent.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(prefix=".bioidea_export_", suffix=".xlsx", dir=str(out_path.parent))
    import os
    os.close(fd)
    wb.save(tmp)
    shutil.move(tmp, out_path)
    return out_path
